# Take memory offer from client as input
# Output formatted memory offer

from openpyxl import Workbook, load_workbook, cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import date
import sys

#
# Functions
#

# Get data from user
def getInputFileName():
    inputFile =  sys.argv[1]
    return inputFile

# Print a lil line
def printLine():
    for i in range(1,25):
        print("-", end="")
    print("")

# Get output file name from user
def getOutputFileName():
    outputFile = sys.argv[2]

    #print("Auto text replacements:")
    #print("%td  - insert todays date seperated by \".\"")
    #outputFile = input("Please input name of output file: ")

    #outputFile = outputFile.replace("%td", date.today().strftime("%m.%d.%y"))
    #if outputFile[len(outputFile) - 5:len(outputFile)] != ".xlsx":
    #    outputFile = outputFile + ".xlsx"
    return outputFile

# Copy the values of one column to another
def copy(source, dest, sourceColumn, destColumn):
    for i in range(2, source.max_row + 1):
        dest.cell(row=i, column=destColumn).value = source.cell(row=i, column=sourceColumn).value

# Find a description given a part number
def getDesc(PrtNmbr, source):
    for i in range(2, source.max_row + 1):
        if str(source.cell(row=i, column=2).value) == PrtNmbr:
            return str(source.cell(row=i,column=3).value)
    return "Not found"
#########################
#                       #
#     Main function     #
#                       #
#########################

#
# Get input from user
#

inputFile = getInputFileName()      # Get name of input file
printLine()                         # Print line
outputFile = getOutputFileName()    # Get name of ouput file

#
# Get workbook
#

wb = load_workbook(inputFile)   # Load work book from input file
ws = wb.active                  # Load active sheet from work book

# TODO: Add in searching the heading for needed parts, and alias for the needed sections
#firstRow = [cell.value for cell in sheet[1]]    # Get first row

#
# Make the output workbook
#

outputWB = Workbook()
outputWS = outputWB.active
outputWS.title = "OutputWorksheet"

#
# Cut out empty lines, and extra headings
#

for row_idx in range(ws.max_row, 1, -1):
    if all(ws.cell(row=row_idx, column=col_idx).value is None for col_idx in range(3, ws.max_column + 1)):
        ws.delete_rows(row_idx)
    if str(ws.cell(row=row_idx, column=1).value) == "Data Class":
        ws.delete_rows(row_idx)

#############################
#                           #
#   Test offer data sheet   #
#                           #
#############################

# Input:    Data Class, Type, Size, Manufacturer, ECC, Rank, Speed, Module #, Qty, Cost, Total
# Output:   Mnfr, Part Number, Description, Qty, Offer
# Descrpt:  Size Speed SpeedMHZ ECC CL## {Pin count} Type

#############################################
#                                           #
#   Copy the new sheet to the output sheet  #
#                                           #
#############################################

copy(ws, outputWS, 4, 1)    # First column  = Mnfr
copy(ws, outputWS, 8, 2)    # Second column = Part Number

#########################
#   Get description     #
#########################

#for i in range(2, ws.max_row + 1):
#    size    =   str(ws.cell(row=i, column=3).value) # Get size
#    speed   =   str(ws.cell(row=i, column=7).value) # Get speed
#    # Find out how to get speedMHZ
#    ECC     =   str(ws.cell(row=i, column=5).value) # Get ECC
#    ECC     =   ECC.replace("ECC/REG", "ECC Registered")        # Fix registered
#    ECC     =   ECC.replace("NON/NON", "non-ECC Unbuffered")    # Fix non-registered
#    # Find out what cl## means
#    # Find pin count
#    sType   =   str(ws.cell(row=i, column=2).value) # Get Slot type

#
# Load Database
#

DBWB = load_workbook("/var/www/xlsxapp/DataBase.xlsx")
Database = DBWB.active
#
# Loop through database checking part numbers
#

for i in range(2, ws.max_row + 1):
    partnumber = str(outputWS.cell(row=i, column=2).value)
    description = getDesc(partnumber, Database)

    #for j in range(2, Database.max_row + 1):
    #    if str(Database.cell(row=j, column=2).value) == partnumber:
    #        description = str(Database.cell(row=j, column=3).value)

    outputWS.cell(row=i, column=3).value = description

DBWB.close()

copy(ws, outputWS, 9, 4)    # Fourth column = Qty

#########################
#   Add in the title    #
#   Format the title    #
#########################

# Put the title back
title = ["Mnfr", "Part Number", "Description", "Qty", "Offer"]
for col_num, value in enumerate(title, start=1):
    outputWS.cell(row=1, column=col_num, value=value)

    border      = Side(border_style="thin", color="000000")
    fill        = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    alignment   = Alignment(horizontal='center')
    font        = Font(name='Apotos Narrow', size=12, bold=False, color='000000')

# Find widest cell and set all cells to that +2
for column in outputWS.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)

    for cell in column:
        # Cell formatting
        cell.border = Border(top=border, left=border, right=border, bottom=border)
        cell.fill = fill
        cell.alignment = alignment
        cell.font = font
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    
    adjusted_width = (max_length + 2)
    outputWS.column_dimensions[column_letter].width = adjusted_width

for cell in outputWS[1]:
    cell.fill = PatternFill(start_color='909090', end_color='909090', fill_type='solid')
    cell.font = Font(name='Apotos Narrow', size=12, bold=True, color='000000')

outputWB.save(str(outputFile))
